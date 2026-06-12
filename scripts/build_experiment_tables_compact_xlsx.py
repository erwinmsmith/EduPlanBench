from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_ORDER = [
    ("Track1_Main", "T1 Main"),
    ("Track1_Specific", "T1 Metrics"),
    ("Track1_Difficulty", "T1 Diff"),
    ("Track2_Main", "T2 Main"),
    ("Track2_PathQuality", "T2 Path"),
    ("Track2_TaskTypes", "T2 Types"),
    ("Track3_Main", "T3 Main"),
    ("Track3_ClosedLoop", "T3 Learn"),
    ("Track3_ActionDiag", "T3 Actions"),
    ("Robustness", "Robust"),
]

HEADER_MAP = {
    "Agent": "Agent",
    "Overall ↑": "Overall",
    "Core ↑": "Core",
    "Track ↑": "Track",
    "GSR ↑": "GSR",
    "PR ↑": "PR",
    "Steps ↓": "Steps",
    "Valid ↑": "Valid",
    "Context ↓": "CtxTok",
    "Time ↓": "Time",
    "Misconception Acc ↑": "MAcc",
    "Feedback Grounding ↑": "Ground",
    "Remediation Match ↑": "Remed",
    "Hint Helpfulness ↑": "Hint",
    "Error-aware Replan ↑": "Replan",
    "Redundancy ↓": "Redund",
    "Direct-answer Rate ↓": "Direct",
    "Easy GSR ↑": "Easy GSR",
    "Easy PR ↑": "Easy PR",
    "Medium GSR ↑": "Med GSR",
    "Medium PR ↑": "Med PR",
    "Hard GSR ↑": "Hard GSR",
    "Hard PR ↑": "Hard PR",
    "Prereq Violation ↓": "PrereqV",
    "Sequence Consistency ↑": "Seq",
    "Resource-Concept Match ↑": "ResMatch",
    "Path Coherence ↑": "Path",
    "Constraint Satisfaction ↑": "Const",
    "Difficulty Alignment ↑": "DiffAlign",
    "Plan Drift ↓": "Drift",
    "Goal-to-Path PR ↑": "GoalPR",
    "Adaptive Replan PR ↑": "AdaptPR",
    "Constraint Planning PR ↑": "ConstPR",
    "Long-context Memory PR ↑": "MemPR",
    "Retention Planning PR ↑": "RetPR",
    "Mastery Gain ↑": "MGain",
    "Retention Gain ↑": "RGain",
    "Learning Efficiency ↑": "Eff",
    "Dropout Risk ↓": "Drop",
    "Overload Rate ↓": "Over",
    "Recovery Rate ↑": "Recover",
    "Simulator Exploit ↓": "Exploit",
    "Exercise %": "Ex%",
    "Review %": "Rev%",
    "Explanation %": "Exp%",
    "Diagnostic %": "Diag%",
    "Avg Difficulty": "AvgDiff",
    "Target-concept Hit ↑": "Hit",
    "Fallback Rate ↓": "Fallback",
    "Unique Resources ↑": "Unique",
    "Track": "Track",
    "H=10 GSR ↑": "H10 GSR",
    "H=10 PR ↑": "H10 PR",
    "H=30 GSR ↑": "H30 GSR",
    "H=30 PR ↑": "H30 PR",
    "H=50 GSR ↑": "H50 GSR",
    "H=50 PR ↑": "H50 PR",
    "H=100 GSR ↑": "H100 GSR",
    "H=100 PR ↑": "H100 PR",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Build compact EduPlanBench Excel tables with Python/openpyxl.")
    parser.add_argument(
        "--experiment-dir",
        type=Path,
        default=Path("outputs/runs/experiment-20260612-210605"),
        help="Experiment directory containing tables/tables.json.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("outputs/workbooks/EduPlanBench_Experiment_Tables_compact.xlsx"),
        help="Output .xlsx path.",
    )
    args = parser.parse_args()

    build_compact_workbook(args.experiment_dir, args.output)


def build_compact_workbook(experiment_dir: Path, output: Path) -> Path:
    tables_path = experiment_dir / "tables" / "tables.json"
    if not tables_path.exists():
        raise FileNotFoundError(f"missing tables JSON: {tables_path}")

    tables = json.loads(tables_path.read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.remove(workbook.active)

    for table_name, sheet_name in SHEET_ORDER:
        rows = tables.get(table_name) or []
        worksheet = workbook.create_sheet(sheet_name)
        _write_sheet(worksheet, rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(output)
    return output


def _write_sheet(worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        worksheet.append(["No data"])
        _style_sheet(worksheet, ["No data"])
        return

    raw_headers = list(rows[0].keys())
    headers = [HEADER_MAP.get(header, header) for header in raw_headers]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([_format_value(row.get(header)) for header in raw_headers])

    _style_sheet(worksheet, headers)


def _format_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 4)
    return value


def _style_sheet(worksheet, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Arial", size=9, bold=True, color="000000")
    body_font = Font(name="Arial", size=9, color="000000")
    thin = Side(style="thin", color="C8D2E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row == 1 else body_font
            cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "center", vertical="center", wrap_text=False)
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill

    worksheet.row_dimensions[1].height = 16
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if idx == 1:
            width = min(24, max(16, _max_text_len(worksheet, idx) + 2))
        else:
            width = min(11, max(7, len(str(header)) + 1))
        worksheet.column_dimensions[letter].width = width

    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 15
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True


def _max_text_len(worksheet, column: int) -> int:
    max_len = 0
    for row_idx in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_idx, column=column).value
        max_len = max(max_len, len(str(value or "")))
    return max_len


if __name__ == "__main__":
    main()
